import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import Button, Entry, Frame, Label, Spinbox, messagebox
from datetime import datetime
import os.path
import pandas as pd
from tkinter.ttk import *
from order.order_win2 import settle_win

price_bev = {'coffee':3500, 'latte':5500, 'smoothie':7000, 'tea':5000}
price_des = {'cake':8000, 'cookie':1500, 'ice-cream':2000, 'tart':4400}
order_total = []
order_bev = {'coffee':0, 'latte':0, 'smoothie':0, 'tea':0}
order_des = {'cake':0, 'cookie':0, 'ice-cream':0, 'tart':0}
# source = {'날짜', '이름', '커피', '라떼', '스무디', '차', '케이크', '쿠키', '아이스크림', '타르트', '총금액'}
sum = 0
sum_price = 0
new_filename = 'C:\\Users\\lkjun\\OneDrive\\바탕 화면\\PythonWorkspace\\python_study\\Python_practice\\order_program.xlsx'


def make_xlsx():
    global new_filename
    
    make_wb = openpyxl.Workbook()
    make_wb.active.title = "order"
    make_wb.save(new_filename)


def clear():
    global sum, order_total, order_bev, order_des, textarea, ent1, ent2
    textarea.delete('1.0', tk.END)
    label5['text'] = "금액 : 0원"
    sum = 0
    order_total = []
    order_bev = {'coffee' : 0, 'latte' : 0, 'smoothie' : 0, 'tea' : 0}
    order_des = {'cake':0, 'cookie':0, 'ice-cream':0, 'tart':0}
    ent1.delete('0', tk.END)
    ent2.delete('0', tk.END)
    ent1.focus()


def add1(item1):
    global sum

    if item1 not in price_bev:
        print("No Drink")
    this_price1 = price_bev.get(item1)
    sum += this_price1
    order_total.append(item1)
    order_bev[item1] += 1
    textarea.insert(tk.INSERT, item1 +" ")
    label5['text'] = "금액 : " + str(sum) + "원"
    # print(type(sum), type(this_price1))


def add2(item2):
    global sum

    if item2 not in price_des:
        print("No Dessert")
    this_price2 = price_des.get(item2)
    sum += this_price2
    order_total.append(item2)
    order_des[item2] += 1
    textarea.insert(tk.INSERT, item2 +" ")
    label5['text'] = "금액 : " + str(sum) + "원"
    # print(type(sum), type(this_price2))


def send():
    global order_total, order_bev, order_des, sum, ent1, ent2
    name = str(ent1.get())
    hp = str(ent2.get())
    print(name, hp)
    print(order_total)
    print(order_bev)
    print(order_des)
    print(str(sum) + "원")
    now_dt = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    wb1 = load_workbook("order_program.xlsx")
    ws1 = wb1['order']
    ws1.append([now_dt, name, hp, 
    order_bev['coffee'], order_bev['latte'], order_bev['smoothie'], order_bev['tea'], 
    order_des['cake'], order_des['cookie'], order_des['ice-cream'], order_des['tart'], 
    sum])
    wb1.save("order_program.xlsx")
    clear()


def maker_one():
    wb2 = load_workbook("order_program.xlsx")
    ws2 = wb2['order']
    ws2.append(["날짜", "이름", "휴대폰 번호", "커피", "라떼", "스무디", "차", "케이크", "쿠키", "아이스크림", "타르트", "총금액"])
    wb2.save("order_program.xlsx")


def but_exit():
    name = str(ent1.get())
    hp = str(ent2.get())
    if(name == ""):
        tk.messagebox.showerror("확인!", "이름을 입력해주세요!!")
        ent1.focus()
        return
    if(hp == ""):
        tk.messagebox.showerror("확인!", "휴대폰번호를 입력해주세요!!")
        ent2.focus()
        return

    msgbox = tk.messagebox.askquestion('확인!', "주문완료하시겠습니까??")
    if msgbox == 'yes':
        if os.path.isfile(new_filename):
            send()
        else:
            make_xlsx()
            maker_one()
            send()


def settle():
    msgbox2 = tk.messagebox.askquestion('정산', "정산 확인하시겠습니까??")
    if msgbox2 == 'yes':
        print("정산")
        settle_win()


def read_xlsx():
    df_rx = pd.read_excel("C:\\Users\\lkjun\\OneDrive\\바탕 화면\\PythonWorkspace\\python_study\\Python_practice\\order_program.xlsx",
    sheet_name = "order")
    df_rx.to_excel("C:\\Users\\lkjun\\OneDrive\\바탕 화면\\PythonWorkspace\\python_study\\Python_practice\\order_program.xlsx"
    , sheet_name = "order", index=False)
    # print(df_rx)


win1 = tk.Tk()

win1.geometry("500x600")

win1.title("Order Program")

frame1 = tk.Frame(win1)
frame1.pack()
frame2 = tk.Frame(win1)
frame2.pack()
frame3 = tk.Frame(win1)
frame3.pack()
frame4 = tk.Frame(win1)
frame4.pack()
frame5 = tk.Frame(win1)
frame5.pack()

label1 = tk.Label(text = "음료", width = 10, height = 2)
label1.place(x = 0, y = 0)
# frame1
label2 = tk.Label(text = "디저트", width = 10, height = 0)
label2.place(x = 0, y = 100)
# frame2

but_1 = tk.Button(frame2, text = "커피", command = lambda : add1('coffee'), width = 10, height = 2)
but_2 = tk.Button(frame2, text = "라떼", command = lambda : add1('latte'), width = 10, height = 2)
but_3 = tk.Button(frame2, text = "스무디", command = lambda : add1('smoothie'), width = 10, height = 2)
but_4 = tk.Button(frame2, text = "차", command = lambda : add1('tea'), width = 10, height = 2)
but_5 = tk.Button(frame4, text = "케이크", command = lambda : add2('cake'), width = 10, height = 2)
but_6 = tk.Button(frame4, text = "쿠키", command = lambda : add2('cookie'), width = 10, height = 2)
but_7 = tk.Button(frame4, text = "아이스크림", command = lambda : add2('ice-cream'), width = 10, height = 2)
but_8 = tk.Button(frame4, text = "타르트", command = lambda : add2('tart'), width = 10, height = 2)
but_9 = tk.Button(frame4, text = "주문완료", command = but_exit, width = 10, height = 2)
but_10 = tk.Button(frame4, text = "정산", command = settle, width = 10, height = 2)

but_1.grid(row=0, column = 0, padx = 20, pady = 40)
but_2.grid(row=0, column = 1, padx = 20, pady = 10)
but_3.grid(row=0, column = 2, padx = 20, pady = 10)
but_4.grid(row=0, column = 3, padx = 20, pady = 10)
but_5.grid(row=3, column = 0, padx = 0, pady = 0)
but_6.grid(row=3, column = 1, padx = 20, pady = 10)
but_7.grid(row=3, column = 2, padx = 20, pady = 10)
but_8.grid(row=3, column = 3, padx = 20, pady = 10)
but_9.grid(row=5, column = 0, padx = 20, pady = 20)
but_10.grid(row=5, column = 3, padx = 20, pady = 20)

label3 = tk.Label(frame5, text = "이름", width=10, height=2).grid(row=0, column=0)
label4 = tk.Label(frame5, text = "휴대폰번호", width=10, height=2).grid(row=1, column=0)

ent1 = tk.Entry(frame5)
ent2 = tk.Entry(frame5)
ent1.grid(row=0, column=1)
ent2.grid(row=1, column=1)

label5 = tk.Label(win1, text="금액 : 0원", width = 100, height = 2, fg="blue")
label5.pack()

textarea = tk.Text(win1)
textarea.pack(padx = 10, pady = 10)

win1.mainloop()