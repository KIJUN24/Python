from openpyxl import load_workbook


file1 = './order_program.xlsx'
i = 1
ws = load_workbook(file1)
wb = ws.active

idx_coffee = []
idx_latte = []
idx_smoothie = []
idx_tea = []
idx_cake = []
idx_cookie = []
idx_icecream = []
idx_tart = []
idx_money = []


while True:
    i += 1
    col_coffee = wb.cell(row=i,column=4).value
    col_latte = wb.cell(row=i,column=5).value
    col_smoothie = wb.cell(row=i,column=6).value
    col_tea = wb.cell(row=i,column=7).value
    col_cake = wb.cell(row=i,column=8).value
    col_cookie = wb.cell(row=i,column=9).value
    col_icecream = wb.cell(row=i,column=10).value
    col_tart = wb.cell(row=i,column=11).value
    col_moeny = wb.cell(row=i,column=12).value


    # print(col1)
    idx_coffee.append(col_coffee)
    idx_latte.append(col_latte)
    idx_smoothie.append(col_smoothie)
    idx_tea.append(col_tea)
    idx_cake.append(col_cake)
    idx_cookie.append(col_cookie)
    idx_icecream.append(col_icecream)
    idx_tart.append(col_tart)
    idx_money.append(col_moeny)

    # print(type(idx.append(col1)))
    if col_coffee is None:
        print(idx_coffee)
        idx_coffee.remove(None)
        idx_latte.remove(None)
        idx_smoothie.remove(None)
        idx_tea.remove(None)
        idx_cake.remove(None)
        idx_cookie.remove(None)
        idx_icecream.remove(None)
        idx_tart.remove(None)
        idx_money.remove(None)
        print(idx_coffee)
        break

print(sum(idx_coffee))
print(sum(idx_latte))
print(sum(idx_smoothie))
print(sum(idx_tea))
print(sum(idx_cake))
print(sum(idx_cookie))
print(sum(idx_icecream))
print(sum(idx_tart))
print(sum(idx_money))

# print(len(idx))


# print(type(i))
# print(type(col1))