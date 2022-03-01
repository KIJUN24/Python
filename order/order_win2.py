import tkinter as tk
from tkinter import Button, Entry, Frame, Label, Spinbox, messagebox
from tkinter.ttk import *
import pandas as pd
from openpyxl import load_workbook

file1 = './order_program.xlsx'
ws = load_workbook(file1)
wb = ws.active
idx_coffee = []
idx_latte = []
idx_smoothie = []
idx_tea = []
idx_cake = []
idx_cookie = []
idx_icecream = []
idx_tart = []
idx_money = []
i = 1

while True:
    i += 1
    col_coffee = wb.cell(row=i,column=4).value
    col_latte = wb.cell(row=i,column=5).value
    col_smoothie = wb.cell(row=i,column=6).value
    col_tea = wb.cell(row=i,column=7).value
    col_cake = wb.cell(row=i,column=8).value
    col_cookie = wb.cell(row=i,column=9).value
    col_icecream = wb.cell(row=i,column=10).value
    col_tart = wb.cell(row=i,column=11).value
    col_moeny = wb.cell(row=i,column=12).value

    idx_coffee.append(col_coffee)
    idx_latte.append(col_latte)
    idx_smoothie.append(col_smoothie)
    idx_tea.append(col_tea)
    idx_cake.append(col_cake)
    idx_cookie.append(col_cookie)
    idx_icecream.append(col_icecream)
    idx_tart.append(col_tart)
    idx_money.append(col_moeny)

# print(type(idx.append(col1)))
    if col_coffee is None:
        idx_coffee.remove(None)
        idx_latte.remove(None)
        idx_smoothie.remove(None)
        idx_tea.remove(None)
        idx_cake.remove(None)
        idx_cookie.remove(None)
        idx_icecream.remove(None)
        idx_tart.remove(None)
        idx_money.remove(None)
        break



def settle_win():

    win2 = tk.Tk()
    win2.geometry("600x700+100+100")
    win2.title("settlemnt window")
    win2.option_add("*Font", "Arial 13")
    win2.resizable(False, False)


    frame_sw1 = tk.Frame(win2)
    frame_sw1.pack()
    frame_sw2 = tk.Frame(win2)
    frame_sw2.pack()
    frame_sw3 = tk.Frame(win2)
    frame_sw3.pack()
    frame_sw4 = tk.Frame(win2, relief="solid")
    frame_sw4.pack()

    label_sw1 = tk.Label(frame_sw1)
    label_sw1.pack()
    label_sw2 = tk.Label(frame_sw1, text = "정산 선택")
    label_sw2.pack()
    label_sw3 = tk.Label(frame_sw4)
    label_sw3.pack()

    def click():

        df1 = pd.read_excel(io = 'C:\\Users\\lkjun\\OneDrive\\바탕 화면\\PythonWorkspace\\python_study\\Python_practice\\order_program.xlsx', sheet_name='order', usecols='D:L')
        df2 = pd.read_excel(io = 'C:\\Users\\lkjun\\OneDrive\\바탕 화면\\PythonWorkspace\\python_study\\Python_practice\\order_program.xlsx', sheet_name='order', usecols='D:G')
        df3 = pd.read_excel(io = 'C:\\Users\\lkjun\\OneDrive\\바탕 화면\\PythonWorkspace\\python_study\\Python_practice\\order_program.xlsx', sheet_name='order', usecols='H:K')
        df4 = pd.read_excel(io = 'C:\\Users\\lkjun\\OneDrive\\바탕 화면\\PythonWorkspace\\python_study\\Python_practice\\order_program.xlsx', sheet_name='order', usecols='L')

        lab_text = cb.get()
        if(lab_text == '총합'):
            label_sw3.config(text = df1)
        elif(lab_text == '음료'):
            label_sw3.config(text = df2)
        elif(lab_text == '디저트'):
            label_sw3.config(text = df3)
        elif(lab_text == '금액_손님수'):
            label_sw3.config(text = df4)
        elif(lab_text == '커피'):
            label_sw3.config(text = sum(idx_coffee))
        elif(lab_text == '라떼'):
            label_sw3.config(text = sum(idx_latte))
        elif(lab_text == '스무디'):
            label_sw3.config(text = sum(idx_smoothie))
        elif(lab_text == '차'):
            label_sw3.config(text = sum(idx_tea))
        elif(lab_text == '케이크'):
            label_sw3.config(text = sum(idx_cake))
        elif(lab_text == '쿠키'):
            label_sw3.config(text = sum(idx_cookie))
        elif(lab_text == '아이스크림'):
            label_sw3.config(text = sum(idx_icecream))
        elif(lab_text == '타르트'):
            label_sw3.config(text = sum(idx_tart))
        elif(lab_text == '총금액'):
            label_sw3.config(text = sum(idx_money))


    # cb_list = ["총합", "음료", "디저트", "총금액"]
    cb_list = ["총합", "음료", "디저트", "금액_손님수", "커피", "라떼", "스무디", "차", "케이크", "쿠키", "아이스크림", "타르트",  "총금액"]
    cb = Combobox(win2)
    cb.config(values = cb_list)
    cb.pack()


    but_sw1 = tk.Button(win2)
    but_sw1.config(text = "선택")
    but_sw1.config(command=click)
    but_sw1.pack()

    # read_xlsx()

    win2.mainloop()


# print(sum(idx_coffee))
# print(sum(idx_latte))
# print(sum(idx_smoothie))
# print(sum(idx_tea))
# print(sum(idx_cake))
# print(sum(idx_cookie))
# print(sum(idx_icecream))
# print(sum(idx_tart))
# print(sum(idx_money))