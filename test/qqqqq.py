from openpyxl import load_workbook

file1 = './exfile.xlsx'

ws = load_workbook(file1)
wb = ws.active

idx = []
for m in range(1,10):
    col1 = wb.cell(row=m+1,column=1).value
    print(col1)
    idx.append(col1)
    # print(idx.append(col1))
    # print(type(idx.append(col1)))
    

# print(type(m))
# print(type(col1))
# print(sum(idx))
# print(len(idx))